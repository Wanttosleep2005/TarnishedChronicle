import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / 'exports' / 'gamersky-armor-acquisition.csv'
XLSX_PATH = ROOT / 'exports' / 'gamersky-armor-acquisition.xlsx'


with CSV_PATH.open('r', encoding='utf-8-sig', newline='') as csv_file:
    reader = csv.DictReader(csv_file)
    columns = reader.fieldnames or []
    rows = list(reader)

workbook = Workbook()
worksheet = workbook.active
worksheet.title = '防具获取攻略'
worksheet.freeze_panes = 'A2'
worksheet.sheet_view.showGridLines = False

worksheet.append(columns)
for row in rows:
    worksheet.append([row.get(column, '') for column in columns])

header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
header_font = Font(color='FFFFFF', bold=True)
body_alignment = Alignment(vertical='top', wrap_text=True)
for cell in worksheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

column_widths = {
    '物品ID': 13,
    '防具名称': 24,
    '部位': 10,
    '获取文本': 66,
    '来源类型': 21,
    '核对方式': 18,
    '攻略分页': 52,
    '来源标题': 42,
}
for column_index, column in enumerate(columns, start=1):
    worksheet.column_dimensions[chr(64 + column_index)].width = column_widths.get(column, 20)

for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = body_alignment
    source_url = row[columns.index('攻略分页')]
    if source_url.value:
        source_url.hyperlink = source_url.value
        source_url.style = 'Hyperlink'

worksheet.auto_filter.ref = worksheet.dimensions
worksheet.row_dimensions[1].height = 28
worksheet.print_title_rows = '1:1'
worksheet.page_setup.orientation = 'landscape'
worksheet.page_setup.fitToWidth = 1
worksheet.sheet_properties.pageSetUpPr.fitToPage = True

workbook.save(XLSX_PATH)
print(f'已导出 {len(rows)} 条防具获取文本：{XLSX_PATH}')
